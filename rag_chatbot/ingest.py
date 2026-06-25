"""Document ingestion and indexing for the local RAG chatbot.

Scans the Docs folder recursively, parses supported file types, chunks text,
embeds chunks with HuggingFace, and builds fresh FAISS and BM25 indexes on each run.
"""

import json
import os
import pickle
import warnings
from pathlib import Path

import faiss
import numpy as np
import openpyxl
import pandas as pd
from langchain_community.document_loaders import (
    BSHTMLLoader,
    Docx2txtLoader,
    JSONLoader,
    PyPDFLoader,
    TextLoader,
)
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_huggingface import HuggingFaceEmbeddings
from rank_bm25 import BM25Okapi

from config import (
    BM25_PICKLE_PATH,
    CHUNK_OVERLAP,
    CHUNK_SIZE,
    DOCUMENT_FOLDER,
    EMBEDDING_MODEL,
    FAISS_INDEX_PATH,
    FAISS_META_PATH,
)

EXTENSION_MAP = {
    ".pdf": "pdf",
    ".docx": "docx",
    ".txt": "text",
    ".md": "text",
    ".html": "html",
    ".htm": "html",
    ".xlsx": "excel",
    ".xls": "excel",
    ".csv": "csv",
    ".json": "json",
}


def ensure_store_dir() -> None:
    store_dir = os.path.dirname(FAISS_INDEX_PATH)
    if store_dir:
        os.makedirs(store_dir, exist_ok=True)


def serialize_row(headers: list[str], row: tuple, sheet_prefix: str = "") -> str:
    lines = []
    for col_name, value in zip(headers, row):
        if value is not None:
            lines.append(f"{col_name}: {value}")
    body = "\n".join(lines)
    if sheet_prefix:
        return f"Sheet: {sheet_prefix}\n{body}"
    return body


def load_excel_chunks(file_path: str) -> list[dict]:
    chunks = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [
            str(cell) if cell is not None else f"Column{i}"
            for i, cell in enumerate(rows[0])
        ]
        for row in rows[1:]:
            text = serialize_row(headers, row, sheet_prefix=sheet_name)
            if text.strip():
                chunks.append({"text": text, "source": file_path, "page": sheet_name})
    wb.close()
    return chunks


def load_csv_chunks(file_path: str) -> list[dict]:
    chunks = []
    df = pd.read_csv(file_path)
    headers = [str(col) for col in df.columns]
    for _, row in df.iterrows():
        row_tuple = tuple(row.values)
        text = serialize_row(headers, row_tuple)
        if text.strip():
            chunks.append({"text": text, "source": file_path, "page": None})
    return chunks


def split_documents(documents: list, source: str) -> list[dict]:
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=CHUNK_SIZE,
        chunk_overlap=CHUNK_OVERLAP,
    )
    chunks = []
    for doc in documents:
        page = doc.metadata.get("page", doc.metadata.get("page_number", None))
        for split_text in splitter.split_text(doc.page_content):
            if split_text.strip():
                chunks.append({"text": split_text, "source": source, "page": page})
    return chunks


def load_file(file_path: str) -> list[dict]:
    ext = Path(file_path).suffix.lower()
    file_type = EXTENSION_MAP.get(ext)

    if file_type == "pdf":
        loader = PyPDFLoader(file_path)
        return split_documents(loader.load(), source=file_path)

    if file_type == "docx":
        loader = Docx2txtLoader(file_path)
        return split_documents(loader.load(), source=file_path)

    if file_type == "text":
        loader = TextLoader(file_path, encoding="utf-8")
        return split_documents(loader.load(), source=file_path)

    if file_type == "html":
        loader = BSHTMLLoader(file_path)
        return split_documents(loader.load(), source=file_path)

    if file_type == "excel":
        return load_excel_chunks(file_path)

    if file_type == "csv":
        return load_csv_chunks(file_path)

    if file_type == "json":
        loader = JSONLoader(file_path, jq_schema=".", text_content=False)
        return split_documents(loader.load(), source=file_path)

    return []


def scan_files(folder: str) -> list[str]:
    paths = []
    for root, _, files in os.walk(folder):
        for name in files:
            paths.append(os.path.join(root, name))
    return sorted(paths)


def ingest() -> None:
    ensure_store_dir()

    if not os.path.isdir(DOCUMENT_FOLDER):
        os.makedirs(DOCUMENT_FOLDER, exist_ok=True)
        print(f"Created empty document folder: {DOCUMENT_FOLDER}")
        print("Add documents and run ingest again.")
        return

    all_files = scan_files(DOCUMENT_FOLDER)
    all_chunks: list[dict] = []
    files_processed = 0

    for file_path in all_files:
        ext = Path(file_path).suffix.lower()
        if ext not in EXTENSION_MAP:
            warnings.warn(f"Skipping unrecognised extension: {file_path}")
            continue

        try:
            chunks = load_file(file_path)
            all_chunks.extend(chunks)
            files_processed += 1
            print(f"Processed: {file_path} ({len(chunks)} chunks)")
        except Exception as exc:
            warnings.warn(f"Failed to process {file_path}: {exc}")

    if not all_chunks:
        print("No chunks created. Add documents to Docs/ and run again.")
        return

    print(f"\nEmbedding {len(all_chunks)} chunks...")
    embeddings_model = HuggingFaceEmbeddings(model_name=EMBEDDING_MODEL)
    texts = [chunk["text"] for chunk in all_chunks]
    vectors = embeddings_model.embed_documents(texts)
    vectors_np = np.array(vectors, dtype=np.float32)

    dim = vectors_np.shape[1]
    index = faiss.IndexFlatL2(dim)
    index.add(vectors_np)
    faiss.write_index(index, FAISS_INDEX_PATH)

    with open(FAISS_META_PATH, "w", encoding="utf-8") as f:
        json.dump(all_chunks, f, ensure_ascii=False, indent=2)

    tokenised_corpus = [text.lower().split() for text in texts]
    bm25_obj = BM25Okapi(tokenised_corpus)
    with open(BM25_PICKLE_PATH, "wb") as f:
        pickle.dump((bm25_obj, all_chunks), f)

    print("\n--- Ingestion summary ---")
    print(f"Files processed : {files_processed}")
    print(f"Total chunks    : {len(all_chunks)}")
    print(f"FAISS vectors   : {index.ntotal}")
    print(f"BM25 corpus     : {len(tokenised_corpus)} documents")
    print(f"Index saved to  : {os.path.dirname(FAISS_INDEX_PATH)}/")


if __name__ == "__main__":
    ingest()
